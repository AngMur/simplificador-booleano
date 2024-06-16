import openpyxl
from openpyxl.styles import Font, Color, PatternFill, Border, Side, Alignment

'''------------------------QMCK------------------------------'''

r = []

'''
Regresa un string que equivale al binario del decimal ingresado
'''
def convertBinary(dec):
  return str(dec) if dec == 1 or dec == 0 else convertBinary(int(dec/2)) + str(int(dec%2))


'''
Rellenea los espacios vacíos con zeros en caso de ser necesario,
"n" es la cantidad de espacios requeridos.
'''
def fill_zeros(bin, n):
  return f"{bin:0>{n}}"


'''
Verifica si los t1 y t2 cambian en una sola posición e indica cual,
en caso de no encontra ningún cambio o encontrar más de uno regresara
False y -1
'''
def find_changes(t1, t2):
  count, index = 0, -1

  for i in range(len(t1)):
    if(t1[i] != t2[i]):
      count += 1
      index = i

  if count == 1:
    return True, index
  else:
    return False, -1

'''
Cambia un caracter en determinada posicion
'''
def replace_index(string, index, new):
  output = ""
  for c in range(len(string)):
    if(c == index):
      output += new
    else:
      output += string[c]
  return output

'''
"CUBO 0" Ordena los miniterminos con respecto a la cantidad de unos
que tengan.
'''
def cb0(min_terms, n):
  terms = cb_creater(n)
  for i in min_terms:
    bin = fill_zeros(convertBinary(i), n)
    terms[bin.count("1")][bin] =  [i, False]
  return terms

'''
Crea un cubo vacío
'''
def cb_creater(n):
  return [{} for _ in range(n+1)]

'''
"CUBO N" Compara todos los elementos de el escalon actual con
el escalon siguiente, hasta llegar al último
'''
def cbn(cb_act):
  n = len(cb_act) - 1
  cb_sig = cb_creater(n-1)
  for i in range(n):
    comparator(cb_act[i], cb_act[i+1], cb_sig, i)
  last_r(cb_act[-1])
  return cb_sig

'''Guarda los residuos del ultimo diccionario'''
def last_r(d):
  global r
  for i in d.keys():
    if(not(d.get(i)[1])):
      r.append((i, d.get(i)[0]))

'''
Compara los elementos de un diccionario con otro (TODOS CON TODOS)
'''
def comparator(d1, d2, cb_siguiente, step):
  global r
  for t1 in d1.keys():
    for t2 in d2.keys():
      cp = find_changes(t1,t2)
      if(cp[0]):
        d1.get(t1)[1] = True
        d2.get(t2)[1] = True
        identifier = f"{d1[t1][0]}-{d2[t2][0]}"
        add_nextgen((replace_index(t1, cp[1], "*"), identifier), cb_siguiente, step)
        #Si cambia en una posicion guardar cambio en siguiente cubo
        #Ademas cambiar valores de FALSE A TRUE

    #AQUELLOS TERMINOS QUE NO SE PUDEN COMBINAR SE ALMACENAN EN R
    if(not(d1.get(t1)[1])):
      r.append((t1, d1.get(t1)[0]))



'''
Agrega dentro del cubo y escalon seleccionado el nuevo termino y
su identd1ificador
'''
def add_nextgen(term_identifier, cb, step):
  cb[step][term_identifier[0]] = [term_identifier[1], False]


def print_cb(cb):
  for i in cb:
    print("{}")
    for j in i.items():
      print(j)
    print(" ")
  print("---------------------------------------------")


'''
Verifica si el cubo esta vacío
'''
def is_empty(cb):
  return all(not bool(d) for d in cb)


def QMCK(min_term, n):
  cubos = []
  cb = cb0(min_term, n)
  while not(is_empty(cb)):
    cubos.append(cb)
    cb = cbn(cb)

  #for i in cubos:
  #  print_cb(i)

  return cubos


def idk(str1, bits):
  output = ""
  for i in range(len(str1)):
    if(str1[i] == "0"):
      output += "~" + bits[i]
    elif(str1[i] == "1"):
      output += bits[i]
    else:
      output += ""
  return output.upper()


def get_funtion():
  global r
  global bits
  funcion = "F = "

  for i in range(len(r)):
    if(i == len(r)-1):
      funcion += idk(r[i][0], bits)
    else:
      funcion += idk(r[i][0], bits) + " + "
  return funcion

'''------------------------QMCK------------------------------'''

'''----------------------EJECUCIÓN---------------------------'''


min_term = [0,4,8,3,5,12,7,11,14,15]
# bits = ["a","b","c","d"]
# n = len(bits)

min_term = input("Miniterminos con 1: ").split(",")
#Convierte la entrada en enteros
min_term = [int(i) for i in min_term]

bits = input("Variables de las que depende: ").split(",")
#Obtiene el numero de variables
n = len(bits)


#Obtiene los cubos, la función y el archivo
cubos = QMCK(min_term, n)
f = get_funtion()


'''------------------------EXCEL-----------------------------'''

#Alinea la celda correspondiente con los valores indicados
def alignment(cell, a_horinzontal, a_vertical):
  alignment = Alignment(horizontal=a_horinzontal, vertical=a_vertical)
  sheet[cell].alignment = alignment


#Genera encabezados formateados en la fila especificada
def headers(row):
  #Identificadores
  D = "D" + str(row)
  E = "E" + str(row)
  F = "F" + str(row)
  #Valores
  sheet[D] = "Termino"
  sheet[E] = "Identificador"
  sheet[F] = "Combinar"
  #Estilos de encabezados
  font = Font(name='Calibri', size=11, bold=False, color=Color(rgb="FFFFFF"))
  sheet[D].font = font
  sheet[E].font = font
  sheet[F].font = font
  #Fondo de encabezados
  fill = PatternFill(start_color="0f2f76", end_color="0f2f76", fill_type="solid")
  sheet[D].fill = fill
  sheet[E].fill = fill
  sheet[F].fill = fill
  #Alineación de encabezados
  alignment(D, 'center', 'center')
  alignment(E, 'center', 'center')
  alignment(F, 'center', 'center')


#Genera encabezados formateados en la fila especificada
def cells(row, term, identifier, combine):
  #Identificadores
  D = "D" + str(row)
  E = "E" + str(row)
  F = "F" + str(row)
  #Valores
  sheet[D] = term
  sheet[E] = identifier
  sheet[F] = str(combine)
  #Estilos de celdas
  font = Font(name='Calibri', size=11, bold=False, color=Color(rgb="000000"))
  sheet[D].font = font
  sheet[E].font = font
  sheet[F].font = font
  #Alineación de celdas
  alignment(D, 'center', 'center')
  alignment(E, 'center', 'center')
  alignment(F, 'center', 'center')



#Genera una separación entre dos filas (solo rellena)
def fill(row):
  #Identificadores
  D = "D" + str(row)
  E = "E" + str(row)
  F = "F" + str(row)
  #Fondo de separacion
  fill = PatternFill(start_color="222222", end_color="222222", fill_type="solid")
  sheet[D].fill = fill
  sheet[E].fill = fill
  sheet[F].fill = fill

#Genera un nombre de cubo
def name_cb(n, row):
  #Identificadores
  E = "E" + str(row)
  CB = "Cubo " + str(n)
  #Valores
  sheet[E] = CB
  #Estilos de celdas
  font = Font(name='Calibri', size=11, bold=False, color=Color(rgb="000000"))
  sheet[E].font = font
  #Alineación de celdas
  alignment(E, 'center', 'center')

#Genera la función booleana en el archivo
def boolean_f():
  #Encabezado de funcion
  sheet['H6'] = "Función: "
  #Estilos de encabezado funcion
  font = Font(name='Calibri', size=11, bold=False, color=Color(rgb="FFFFFF"))
  sheet['H6'].font = font
  #Fondo de encabezado funcion
  fill = PatternFill(start_color="0f2f76", end_color="0f2f76", fill_type="solid")
  sheet['H6'].fill = fill
  #Alineación de encabezado funcion
  alignment('H6', 'center', 'center')
  #Función
  sheet['H7'] = f
  #Color
  font = Font(name='Calibri', size=11, bold=True, color=Color(rgb="000000"))
  sheet['H7'].font = font
  #Alineacion
  alignment('H7', 'center', 'center')



# Crea un nuevo libro de Excel
workbook = openpyxl.Workbook()
sheet = workbook.active

#Formatear las columnas
sheet.column_dimensions['D'].number_format = '@'
sheet.column_dimensions['E'].number_format = '@'
sheet.column_dimensions['F'].number_format = '@'
sheet.column_dimensions['H'].number_format = '@'

#Ancho de columnas
sheet.column_dimensions['D'].width = 9.71
sheet.column_dimensions['E'].width = 18.57
sheet.column_dimensions['F'].width = 12.14
sheet.column_dimensions['H'].width = 77.43

#Titulo
sheet.merge_cells('D2:F2')
sheet.row_dimensions[2].height = 28.5
sheet['D2'] = "Simplificación QMCK"
#Estilos de titulo
font = Font(name='Bahnschrift', size=14, bold=True, color=Color(rgb="000000"))
sheet['D2'].font = font
#Alineación titulo
alignment('D2', 'center', 'center')


#Función booleana
boolean_f()



#Recorrido de cubos a documento excel
row = 4
ncb = 0
for cb in cubos:
  name_cb(ncb, row)
  ncb += 1
  row += 2
  headers(row)
  row += 1
  for dc in cb:
    if(dc):
      for i in dc.items():
        cells(row, i[0], i[1][0], i[1][1])
        row += 1
    else:
      cells(row, " "," "," ")
      row += 1
    fill(row)
    row += 1
  row += 3


# Guardar el libro de Excel
workbook.save('simplificaciones.xlsx')

'''------------------------EXCEL-----------------------------'''

'''----------------------EJECUCIÓN TERMINADA---------------------------'''
